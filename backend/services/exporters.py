"""
services/exporters.py — Export campaign data sang GeoJSON / KML / BoQ XLSX.

Tuân thủ GeoJSON RFC 7946 và KML 2.2 spec — KHÔNG bọc vào response wrapper
vì là file format chuẩn quốc tế (đồng thuận với routers/measurements.py).
"""

from __future__ import annotations

import io
import xml.etree.ElementTree as ET
from xml.dom import minidom


# ─────────────────────────────────────────────────────────────
# GeoJSON
# ─────────────────────────────────────────────────────────────

def measurements_to_geojson(rows: list[dict]) -> dict:
    """rows từ DB query (lat, lng, rssi_dbm, ...) → GeoJSON FeatureCollection."""
    return {
        "type": "FeatureCollection",
        "features": [
            {
                "type": "Feature",
                "geometry": {"type": "Point", "coordinates": [r["lng"], r["lat"]]},
                "properties": {
                    "rssiDbm":         r.get("rssi_dbm"),
                    "snrDb":           r.get("snr_db"),
                    "spreadingFactor": r.get("spreading_factor"),
                    "measuredAt":      r["measured_at"].isoformat()
                                       if r.get("measured_at") else None,
                },
            }
            for r in rows
        ],
    }


# ─────────────────────────────────────────────────────────────
# KML 2.2 (Google Earth, Maps.me, QGIS)
# ─────────────────────────────────────────────────────────────

def measurements_to_kml(rows: list[dict], campaign_name: str) -> str:
    """rows → KML XML string với màu theo ngưỡng LoRa Alliance CVT."""
    kml      = ET.Element("kml", xmlns="http://www.opengis.net/kml/2.2")
    document = ET.SubElement(kml, "Document")
    ET.SubElement(document, "name").text = campaign_name

    # Style cho 4 mức RSSI (KML color = AABBGGRR)
    styles = {
        "strong":   "ff2600a5",   # đỏ
        "medium":   "ff61aefd",   # vàng-cam
        "weak":     "ffd1ad74",   # xanh nhạt
        "veryWeak": "ff953631",   # xanh đậm
    }
    for sid, color in styles.items():
        style     = ET.SubElement(document, "Style", id=sid)
        icon_st   = ET.SubElement(style, "IconStyle")
        ET.SubElement(icon_st, "color").text = color
        ET.SubElement(icon_st, "scale").text = "0.8"

    for r in rows:
        rssi = r.get("rssi_dbm")
        if rssi is None:
            continue
        if   rssi >= -90:  sid = "strong"
        elif rssi >= -105: sid = "medium"
        elif rssi >= -120: sid = "weak"
        else:              sid = "veryWeak"

        pm = ET.SubElement(document, "Placemark")
        ET.SubElement(pm, "name").text       = f"{rssi:.0f} dBm"
        ET.SubElement(pm, "styleUrl").text   = f"#{sid}"
        ET.SubElement(pm, "description").text = (
            f"RSSI: {rssi:.1f} dBm\n"
            f"SNR: {r.get('snr_db', 'N/A')} dB\n"
            f"SF: {r.get('spreading_factor', 'N/A')}"
        )
        point = ET.SubElement(pm, "Point")
        ET.SubElement(point, "coordinates").text = f"{r['lng']},{r['lat']},0"

    raw = ET.tostring(kml, encoding="unicode")
    return minidom.parseString(raw).toprettyxml(indent="  ")


# ─────────────────────────────────────────────────────────────
# BoQ XLSX (Bill of Quantities — Persona 7)
# ─────────────────────────────────────────────────────────────

def gateways_to_boq_xlsx(gateways: list[dict], campaign_name: str) -> bytes:
    """
    Bill of Quantities cho danh sách gateway: model, cột buồng, cáp,
    chi phí ước tính (đơn giá đặt mặc định, sales tự sửa lại).
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "BoQ"

    headers = [
        "STT", "Tên gateway", "EUI", "Vĩ độ", "Kinh độ",
        "Độ cao (m)", "Antenna height (m)", "Tx Power (dBm)",
        "Đơn vị", "Số lượng", "Đơn giá (VND)", "Thành tiền (VND)",
    ]

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="534AB7")
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    UNIT_PRICE = 25_000_000   # placeholder — sales sửa lại
    total = 0
    for i, gw in enumerate(gateways, start=1):
        qty       = 1
        line_cost = qty * UNIT_PRICE
        total    += line_cost
        ws.append([
            i,
            gw.get("name") or f"Gateway-{i}",
            gw.get("gatewayEui", ""),
            gw.get("latitude"),
            gw.get("longitude"),
            gw.get("altitudeM"),
            gw.get("antennaHeightM"),
            gw.get("txPowerDbm"),
            "bộ", qty, UNIT_PRICE, line_cost,
        ])

    total_row = ws.max_row + 1
    ws.cell(row=total_row, column=11, value="TỔNG").font   = Font(bold=True)
    ws.cell(row=total_row, column=12, value=total).font    = Font(bold=True)

    # Auto-fit cột (xấp xỉ)
    for col_idx, h in enumerate(headers, start=1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max(len(h) + 2, 14)

    ws.cell(row=ws.max_row + 2, column=1,
            value=f"Báo giá BoQ — Chiến dịch: {campaign_name}").font = Font(italic=True, size=9)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()